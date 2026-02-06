"""
EPFO (Employees' Provident Fund Organisation) passbook parser.

Parses the EPFO e-Passbook PDF / HTML downloaded from:
  - passbook.epfindia.gov.in → Login → Download Passbook

EPFO passbook contains:
  - UAN (Universal Account Number)
  - Member ID
  - Establishment name / employer
  - Monthly contributions (Employee share, Employer share)
  - Interest credited
  - Running balance
  - Pension fund (EPS) contributions

This parser extracts the most recent balance from the passbook.
It also supports the XLSX / CSV download format from some payroll providers.
"""

import csv
import re
from pathlib import Path
from .base import BaseParser, ParseResult, EPFOHolding


class EPFOParser(BaseParser):
    """Parser for EPFO Passbook (PDF / CSV / XLSX)."""

    broker_name = "EPFO"
    file_patterns = [
        "EPFO_*.pdf", "epfo_*.pdf",
        "EPF_Passbook*.pdf", "epf_passbook*.pdf",
        "UAN_Passbook*.pdf", "passbook_*.pdf",
        "EPFO_*.xlsx", "epfo_*.xlsx", "EPFO_*.csv", "epfo_*.csv",
        "EPF_*.xlsx", "EPF_*.csv",
    ]

    def parse(self, filepath: str) -> ParseResult:
        result = ParseResult(broker_name=self.broker_name)
        ext = Path(filepath).suffix.lower()

        try:
            if ext == '.pdf':
                self._parse_pdf(filepath, result)
            elif ext == '.xlsx':
                self._parse_xlsx(filepath, result)
            elif ext == '.csv':
                self._parse_csv(filepath, result)
            else:
                result.errors.append(f"Unsupported file format: {ext}")
        except Exception as e:
            result.errors.append(f"Error parsing EPFO file {filepath}: {e}")

        return result

    # -----------------------------------------------------------------
    # PDF parsing — most common format from passbook.epfindia.gov.in
    # -----------------------------------------------------------------
    def _parse_pdf(self, filepath: str, result: ParseResult):
        text = self._extract_pdf_text(filepath)
        if not text or len(text) < 50:
            result.errors.append("EPFO PDF appears empty or unreadable")
            return

        # Extract UAN
        uan = ""
        uan_match = re.search(r'UAN\s*[:\-]?\s*(\d{12})', text, re.IGNORECASE)
        if uan_match:
            uan = uan_match.group(1)

        # Extract Member ID
        member_id = ""
        mid_patterns = [
            r'Member\s*Id\s*[:\-]?\s*([A-Z]{2}/\d+/\d+)',
            r'Member\s*Id\s*[:\-]?\s*(\w+/\w+/\w+)',
        ]
        for pat in mid_patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                member_id = m.group(1)
                break

        # Extract Establishment name
        establishment = ""
        est_match = re.search(
            r'(?:establishment|employer|company)\s*(?:name)?\s*[:\-]?\s*([A-Z][A-Za-z\s&.,]+?)(?:\n|Member|UAN|\d{2}/)',
            text, re.IGNORECASE
        )
        if est_match:
            establishment = est_match.group(1).strip()

        # Extract balances — look for the closing balance / last row
        # EPFO passbook rows: Date | Particulars | EE Share | ER Share | Pension | Remarks
        # The last row or a "Closing Balance" line has the final amounts

        employee_share = 0.0
        employer_share = 0.0
        pension_share = 0.0

        # Strategy 1: Look for explicit closing balance
        closing_patterns = [
            r'(?:closing|total|net)\s*balance\s*[:\-]?\s*₹?\s*([\d,]+\.?\d*)',
        ]
        total_balance = 0.0
        for pat in closing_patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                total_balance = _parse_amount(m.group(1))
                break

        # Strategy 2: Extract EE / ER / Pension shares from tabular data
        # Common pattern: numbers in 3 columns
        share_patterns = [
            # "Employee Share" ... amount
            r'Employee\s*(?:Share|Contribution)\s*[:\-]?\s*₹?\s*([\d,]+\.?\d*)',
            r'Employer\s*(?:Share|Contribution)\s*[:\-]?\s*₹?\s*([\d,]+\.?\d*)',
            r'Pension\s*(?:Share|Fund|Contribution)\s*[:\-]?\s*₹?\s*([\d,]+\.?\d*)',
        ]
        for i, pat in enumerate(share_patterns):
            # Find ALL matches, take the last (most recent / closing)
            matches = re.findall(pat, text, re.IGNORECASE)
            if matches:
                value = _parse_amount(matches[-1])
                if i == 0:
                    employee_share = value
                elif i == 1:
                    employer_share = value
                else:
                    pension_share = value

        # Strategy 3: Look for running totals in rows — each row has 3 amounts
        # Take the last set of 3 numbers that look like balances
        if employee_share == 0 and total_balance == 0:
            # Find all amount-like numbers
            amount_rows = re.findall(
                r'([\d,]+\.?\d{0,2})\s+([\d,]+\.?\d{0,2})\s+([\d,]+\.?\d{0,2})',
                text
            )
            if amount_rows:
                last_row = amount_rows[-1]
                vals = [_parse_amount(x) for x in last_row]
                if all(v > 100 for v in vals):  # Sanity check
                    employee_share = vals[0]
                    employer_share = vals[1]
                    pension_share = vals[2]

        if total_balance == 0:
            total_balance = employee_share + employer_share + pension_share

        # Extract interest earned
        interest = 0.0
        int_match = re.search(
            r'(?:interest|int\.?)\s*(?:credited|earned)?\s*[:\-]?\s*₹?\s*([\d,]+\.?\d*)',
            text, re.IGNORECASE
        )
        if int_match:
            interest = _parse_amount(int_match.group(1))

        # Statement date
        date_match = re.search(
            r'(?:as\s+on|date|upto|till)\s*[:\-]?\s*(\d{1,2}[\-/]\w{3}[\-/]\d{4}|\d{1,2}[\-/]\d{1,2}[\-/]\d{4})',
            text, re.IGNORECASE
        )
        stmt_date = date_match.group(1) if date_match else ""

        if total_balance > 0 or employee_share > 0:
            holding = EPFOHolding(
                uan=uan,
                member_id=member_id,
                establishment=establishment,
                employee_share=employee_share,
                employer_share=employer_share,
                pension_share=pension_share,
                total_balance=total_balance,
                interest_earned=interest,
                statement_date=stmt_date,
            )
            result.epfo_holdings.append(holding)
        else:
            result.errors.append(
                "Could not extract EPFO balance from PDF. "
                "Try downloading a fresh passbook from passbook.epfindia.gov.in"
            )

    def _extract_pdf_text(self, filepath: str) -> str:
        from PyPDF2 import PdfReader
        reader = PdfReader(filepath)
        parts = []
        for page in reader.pages:
            t = page.extract_text()
            if t:
                parts.append(t)
        return '\n'.join(parts)

    # -----------------------------------------------------------------
    # XLSX parsing — from payroll providers or manual export
    # -----------------------------------------------------------------
    def _parse_xlsx(self, filepath: str, result: ParseResult):
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            result.errors.append("EPFO XLSX file has too few rows")
            return

        self._parse_tabular(rows, result)

    # -----------------------------------------------------------------
    # CSV parsing
    # -----------------------------------------------------------------
    def _parse_csv(self, filepath: str, result: ParseResult):
        rows = []
        for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                with open(filepath, newline='', encoding=enc) as f:
                    reader = csv.reader(f)
                    rows = [r for r in reader]
                break
            except (UnicodeDecodeError, UnicodeError):
                continue

        if len(rows) < 2:
            result.errors.append("EPFO CSV file has too few rows")
            return

        self._parse_tabular(rows, result)

    # -----------------------------------------------------------------
    # Common tabular parser for XLSX / CSV
    # -----------------------------------------------------------------
    def _parse_tabular(self, rows: list, result: ParseResult):
        """Parse EPFO data from tabular rows (XLSX or CSV)."""
        header_idx = -1
        col_map = {}

        COLUMN_HINTS = {
            'uan': ['uan', 'universal account'],
            'member_id': ['member id', 'member_id', 'memberid'],
            'establishment': ['establishment', 'employer', 'company'],
            'employee_share': ['employee share', 'employee contribution', 'ee share', 'ee contribution'],
            'employer_share': ['employer share', 'employer contribution', 'er share', 'er contribution'],
            'pension_share': ['pension', 'eps', 'pension share', 'pension fund'],
            'total_balance': ['total', 'balance', 'closing', 'net'],
            'interest': ['interest', 'int'],
            'date': ['date', 'as on', 'period'],
        }

        for i, row in enumerate(rows):
            cells = [str(c).strip().lower() if c else '' for c in row]
            joined = ' '.join(cells)
            # Look for header row — must contain at least 2 relevant keywords
            matches = 0
            for key, hints in COLUMN_HINTS.items():
                for hint in hints:
                    if hint in joined:
                        matches += 1
                        break
            if matches >= 2:
                header_idx = i
                # Build column map
                for ci, cell in enumerate(cells):
                    for key, hints in COLUMN_HINTS.items():
                        for hint in hints:
                            if hint in cell and key not in col_map:
                                col_map[key] = ci
                                break
                break

        if header_idx < 0:
            result.errors.append("Cannot detect header row in EPFO file")
            return

        # Parse data rows — take last row as most recent
        last_valid_row = None
        for row in rows[header_idx + 1:]:
            cells = [str(c).strip() if c else '' for c in row]
            if not any(cells):
                continue
            last_valid_row = cells

        if not last_valid_row:
            result.errors.append("No data rows found in EPFO file")
            return

        def _get(key: str) -> str:
            idx = col_map.get(key, -1)
            if 0 <= idx < len(last_valid_row):
                return last_valid_row[idx]
            return ''

        ee = _parse_amount(_get('employee_share'))
        er = _parse_amount(_get('employer_share'))
        pension = _parse_amount(_get('pension_share'))
        total = _parse_amount(_get('total_balance'))
        if total == 0:
            total = ee + er + pension

        if total > 0 or ee > 0:
            holding = EPFOHolding(
                uan=_get('uan'),
                member_id=_get('member_id'),
                establishment=_get('establishment'),
                employee_share=ee,
                employer_share=er,
                pension_share=pension,
                total_balance=total,
                interest_earned=_parse_amount(_get('interest')),
                statement_date=_get('date'),
            )
            result.epfo_holdings.append(holding)
        else:
            result.errors.append("Could not extract EPFO balance from file")


def _parse_amount(s: str) -> float:
    """Parse Indian-format amount to float."""
    try:
        return float(str(s).replace(',', '').replace('₹', '').strip())
    except (ValueError, TypeError):
        return 0.0
