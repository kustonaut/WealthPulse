"""
Mutual Funds statement parser.

Parses MF Central / Groww / Kuvera consolidated mutual fund XLSX exports.
Expected format:
  Scheme Name | AMC | Category | Sub Category | Folio Number | Source | Units |
  Invested Value | Current Value | Returns | XIRR
Header row is auto-detected (looks for "Scheme Name").
Automatically consolidates multiple folios of the same fund.
"""

from .base import BaseParser, ParseResult, MutualFundHolding


class MutualFundParser(BaseParser):
    """Parser for consolidated mutual fund XLSX exports."""

    broker_name = "MutualFunds"
    file_patterns = ["MutualFunds_*.xlsx", "mf_*.xlsx", "MF_*.xlsx", "mutual_funds_*.xlsx"]

    def parse(self, filepath: str) -> ParseResult:
        import openpyxl
        result = ParseResult(broker_name=self.broker_name)

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            result.errors.append(f"Cannot open {filepath}: {e}")
            return result

        # Try 'Holdings' sheet first
        if 'Holdings' in wb.sheetnames:
            ws = wb['Holdings']
        else:
            ws = wb[wb.sheetnames[0]]

        # Find header row with 'Scheme Name'
        header_row = None
        for r in range(1, min(30, ws.max_row + 1)):
            val = ws.cell(r, 1).value
            if val and 'scheme name' in str(val).lower():
                header_row = r
                break

        if not header_row:
            result.errors.append("Could not find header row with 'Scheme Name'")
            return result

        # Parse raw rows
        raw_funds = []
        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name or str(name).strip() == '':
                continue

            name = str(name).strip()
            amc = str(ws.cell(r, 2).value or '').strip()
            category = str(ws.cell(r, 3).value or '').strip()
            sub_category = str(ws.cell(r, 4).value or '').strip()
            folio = str(ws.cell(r, 5).value or '').strip()
            source = str(ws.cell(r, 6).value or '').strip()
            units = float(ws.cell(r, 7).value or 0)
            invested = float(ws.cell(r, 8).value or 0)
            current = float(ws.cell(r, 9).value or 0)
            # Returns in col 10, XIRR in col 11
            xirr = float(ws.cell(r, 11).value or 0)

            raw_funds.append(MutualFundHolding(
                name=name, amc=amc, category=category, sub_category=sub_category,
                folio=folio, invested=invested, current=current, xirr=xirr,
                units=units, source=source,
            ))

        # Consolidate multiple folios of the same fund
        consolidated = {}
        for mf in raw_funds:
            if mf.name not in consolidated:
                consolidated[mf.name] = MutualFundHolding(
                    name=mf.name, amc=mf.amc, category=mf.category,
                    sub_category=mf.sub_category, folio=mf.folio,
                    invested=0, current=0, xirr=0, units=0, source=mf.source,
                )
            c = consolidated[mf.name]
            # Weighted XIRR for consolidation
            old_weight = c.invested
            new_weight = mf.invested
            total_weight = old_weight + new_weight
            if total_weight > 0:
                c.xirr = round((c.xirr * old_weight + mf.xirr * new_weight) / total_weight, 2)
            c.invested += mf.invested
            c.current += mf.current
            c.units += mf.units

        result.mutual_funds = list(consolidated.values())
        # Round values
        for mf in result.mutual_funds:
            mf.invested = round(mf.invested, 2)
            mf.current = round(mf.current, 2)

        return result
