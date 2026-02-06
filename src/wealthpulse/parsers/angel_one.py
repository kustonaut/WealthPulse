"""
Angel One (Angel Broking) statement parser.

Parses Angel One equity holdings CSV/XLSX export.
Angel One exports can be downloaded from:
  Angel One App → Portfolio → Holdings → Download

Expected XLSX format:
  Script/Stock | ISIN | Sector | Qty | Avg Cost | LTP | Current Value | P&L | % P&L
  OR (CSV from web):
  Scrip Name, ISIN, Qty, Avg Price, LTP, Cur. Value, P&L, % Change

Header row is auto-detected.
"""

import csv
import re
from .base import BaseParser, ParseResult, StockHolding


class AngelOneParser(BaseParser):
    """Parser for Angel One (Angel Broking) equity holdings export."""

    broker_name = "Angel One"
    file_patterns = [
        "Angel_*.xlsx", "angel_*.xlsx", "AngelOne_*.xlsx",
        "AngelBroking_*.xlsx", "angelone_*.xlsx",
        "Angel_*.csv", "angel_*.csv", "AngelOne_*.csv",
    ]

    def parse(self, filepath: str) -> ParseResult:
        if filepath.lower().endswith('.csv'):
            return self._parse_csv(filepath)
        return self._parse_xlsx(filepath)

    def _parse_xlsx(self, filepath: str) -> ParseResult:
        import openpyxl
        result = ParseResult(broker_name=self.broker_name)

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            result.errors.append(f"Cannot open {filepath}: {e}")
            return result

        ws = wb.active

        # Find header row — look for "Script", "Stock", "Scrip", or "Symbol"
        header_row = None
        col_map = {}
        for r in range(1, min(30, ws.max_row + 1)):
            for c in range(1, min(20, ws.max_column + 1)):
                val = str(ws.cell(r, c).value or '').strip().lower()
                if val in ('script', 'stock', 'scrip', 'scrip name', 'symbol', 'stock name'):
                    header_row = r
                    break
            if header_row:
                break

        if not header_row:
            result.errors.append("Could not find header row in Angel One file")
            return result

        # Map column headers
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(header_row, c).value or '').strip().lower()
            col_map[val] = c

        # Identify column indices with flexible matching
        sym_col = _find_col(col_map, ['script', 'stock', 'scrip', 'scrip name', 'symbol', 'stock name'])
        isin_col = _find_col(col_map, ['isin'])
        qty_col = _find_col(col_map, ['qty', 'quantity', 'net qty'])
        avg_col = _find_col(col_map, ['avg cost', 'avg price', 'average price', 'buy avg'])
        ltp_col = _find_col(col_map, ['ltp', 'last price', 'close price', 'current price', 'closing price'])
        cur_val_col = _find_col(col_map, ['current value', 'cur. value', 'cur value', 'mkt value', 'market value'])
        sector_col = _find_col(col_map, ['sector', 'industry'])

        if not sym_col or not qty_col:
            result.errors.append(f"Could not map required columns. Found headers: {list(col_map.keys())}")
            return result

        for r in range(header_row + 1, ws.max_row + 1):
            symbol_raw = ws.cell(r, sym_col).value
            if not symbol_raw or str(symbol_raw).strip() == '':
                continue

            symbol = self._normalize_symbol(str(symbol_raw).strip().upper())
            # Angel sometimes exports NSE:SYMBOL — strip exchange prefix
            if ':' in symbol:
                symbol = symbol.split(':')[-1]

            isin = str(ws.cell(r, isin_col).value or '') if isin_col else ''
            qty = _safe_float(ws.cell(r, qty_col).value)
            avg_price = _safe_float(ws.cell(r, avg_col).value) if avg_col else 0
            closing_price = _safe_float(ws.cell(r, ltp_col).value) if ltp_col else 0
            current_value = _safe_float(ws.cell(r, cur_val_col).value) if cur_val_col else 0
            sector = str(ws.cell(r, sector_col).value or '') if sector_col else ''

            invested = round(qty * avg_price, 2) if avg_price > 0 else 0
            closing_value = current_value if current_value > 0 else round(qty * closing_price, 2)

            if qty <= 0:
                continue

            result.stocks.append(StockHolding(
                symbol=symbol,
                isin=isin,
                quantity=qty,
                avg_price=round(avg_price, 2),
                invested=invested,
                closing_price=round(closing_price, 2),
                closing_value=round(closing_value, 2),
                broker=self.broker_name,
                sector=sector,
            ))

        return result

    def _parse_csv(self, filepath: str) -> ParseResult:
        result = ParseResult(broker_name=self.broker_name)

        try:
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                headers = {h.strip().lower(): h for h in reader.fieldnames} if reader.fieldnames else {}

                sym_key = _find_csv_key(headers, ['scrip name', 'script', 'stock', 'symbol'])
                qty_key = _find_csv_key(headers, ['qty', 'quantity', 'net qty'])
                avg_key = _find_csv_key(headers, ['avg price', 'avg cost', 'buy avg', 'average price'])
                ltp_key = _find_csv_key(headers, ['ltp', 'last price', 'close price'])
                isin_key = _find_csv_key(headers, ['isin'])

                if not sym_key or not qty_key:
                    result.errors.append(f"Cannot find required CSV columns. Headers: {list(headers.keys())}")
                    return result

                for row in reader:
                    symbol_raw = row.get(headers.get(sym_key, ''), '').strip()
                    if not symbol_raw:
                        continue

                    symbol = self._normalize_symbol(symbol_raw.upper())
                    if ':' in symbol:
                        symbol = symbol.split(':')[-1]

                    qty = _safe_float(row.get(headers.get(qty_key, ''), 0))
                    avg_price = _safe_float(row.get(headers.get(avg_key, ''), 0)) if avg_key else 0
                    closing_price = _safe_float(row.get(headers.get(ltp_key, ''), 0)) if ltp_key else 0
                    isin = row.get(headers.get(isin_key, ''), '') if isin_key else ''

                    if qty <= 0:
                        continue

                    result.stocks.append(StockHolding(
                        symbol=symbol,
                        isin=isin,
                        quantity=qty,
                        avg_price=round(avg_price, 2),
                        invested=round(qty * avg_price, 2),
                        closing_price=round(closing_price, 2),
                        closing_value=round(qty * closing_price, 2),
                        broker=self.broker_name,
                    ))

        except Exception as e:
            result.errors.append(f"Error parsing CSV: {e}")

        return result


def _find_col(col_map: dict, candidates: list[str]) -> int | None:
    """Find column index from a list of candidate header names."""
    for c in candidates:
        if c in col_map:
            return col_map[c]
    return None


def _find_csv_key(headers: dict, candidates: list[str]) -> str | None:
    """Find CSV header key from candidates."""
    for c in candidates:
        if c in headers:
            return c
    return None


def _safe_float(val) -> float:
    """Safely convert a value to float."""
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(',', '').replace('₹', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return 0.0
