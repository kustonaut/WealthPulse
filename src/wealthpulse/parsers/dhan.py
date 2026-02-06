"""
Dhan broker statement parser.

Parses Dhan equity holdings CSV export.
Download: Dhan Web/App → Portfolio → Holdings → Export CSV

Expected format:
  Trading Symbol, ISIN, Exchange, Segment, Qty, Avg. Cost Price, Close Price, Cur. Value, Day P&L, Overall P&L

Header row is auto-detected.
"""

import csv
from .base import BaseParser, ParseResult, StockHolding


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(',', '').replace('₹', '').strip())
    except (ValueError, TypeError):
        return 0.0


class DhanParser(BaseParser):
    """Parser for Dhan equity holdings export (CSV/XLSX)."""

    broker_name = "Dhan"
    file_patterns = [
        "Dhan_*.csv", "dhan_*.csv", "dhan-holdings*.csv",
        "Dhan_*.xlsx", "dhan_*.xlsx", "dhan-holdings*.xlsx",
    ]

    def parse(self, filepath: str) -> ParseResult:
        if filepath.lower().endswith('.csv'):
            return self._parse_csv(filepath)
        return self._parse_xlsx(filepath)

    def _parse_csv(self, filepath: str) -> ParseResult:
        result = ParseResult(broker_name=self.broker_name)

        try:
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    result.errors.append("Empty CSV file")
                    return result

                headers = {h.strip().lower(): h for h in reader.fieldnames}

                for row in reader:
                    symbol_raw = None
                    for key in ['trading symbol', 'symbol', 'scrip', 'stock', 'instrument']:
                        mapped = headers.get(key)
                        if mapped and row.get(mapped):
                            symbol_raw = row[mapped].strip()
                            break

                    if not symbol_raw:
                        continue

                    symbol = self._normalize_symbol(symbol_raw.upper())
                    # Dhan may format as EXCHANGE-SYMBOL
                    if '-' in symbol and symbol.split('-')[0] in ('NSE', 'BSE'):
                        symbol = symbol.split('-', 1)[-1]

                    qty = 0
                    for key in ['qty', 'quantity', 'net qty', 'total qty']:
                        mapped = headers.get(key)
                        if mapped and row.get(mapped):
                            qty = _safe_float(row[mapped])
                            break

                    avg_price = 0
                    for key in ['avg. cost price', 'avg cost', 'avg price', 'buy price', 'cost price']:
                        mapped = headers.get(key)
                        if mapped and row.get(mapped):
                            avg_price = _safe_float(row[mapped])
                            break

                    ltp = 0
                    for key in ['close price', 'ltp', 'last price', 'current price']:
                        mapped = headers.get(key)
                        if mapped and row.get(mapped):
                            ltp = _safe_float(row[mapped])
                            break

                    cur_value = 0
                    for key in ['cur. value', 'current value', 'mkt value', 'market value']:
                        mapped = headers.get(key)
                        if mapped and row.get(mapped):
                            cur_value = _safe_float(row[mapped])
                            break

                    isin = ''
                    for key in ['isin', 'isin code']:
                        mapped = headers.get(key)
                        if mapped and row.get(mapped):
                            isin = row[mapped].strip()
                            break

                    if qty <= 0:
                        continue

                    result.stocks.append(StockHolding(
                        symbol=symbol,
                        isin=isin,
                        quantity=qty,
                        avg_price=round(avg_price, 2),
                        invested=round(qty * avg_price, 2),
                        closing_price=round(ltp, 2),
                        closing_value=cur_value if cur_value > 0 else round(qty * ltp, 2),
                        broker=self.broker_name,
                    ))

        except Exception as e:
            result.errors.append(f"Error parsing Dhan CSV: {e}")

        return result

    def _parse_xlsx(self, filepath: str) -> ParseResult:
        import openpyxl
        result = ParseResult(broker_name=self.broker_name)

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            result.errors.append(f"Cannot open {filepath}: {e}")
            return result

        ws = wb.active

        header_row = None
        col_map = {}
        for r in range(1, min(30, ws.max_row + 1)):
            for c in range(1, min(20, ws.max_column + 1)):
                val = str(ws.cell(r, c).value or '').strip().lower()
                if val in ('trading symbol', 'symbol', 'scrip', 'instrument', 'stock'):
                    header_row = r
                    break
            if header_row:
                break

        if not header_row:
            result.errors.append("Could not find header row in Dhan file")
            return result

        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(header_row, c).value or '').strip().lower()
            col_map[val] = c

        sym_col = col_map.get('trading symbol') or col_map.get('symbol') or col_map.get('scrip')
        isin_col = col_map.get('isin')
        qty_col = col_map.get('qty') or col_map.get('quantity')
        avg_col = col_map.get('avg. cost price') or col_map.get('avg price') or col_map.get('avg cost')
        ltp_col = col_map.get('close price') or col_map.get('ltp') or col_map.get('last price')
        val_col = col_map.get('cur. value') or col_map.get('current value')

        if not sym_col or not qty_col:
            result.errors.append(f"Cannot map columns. Headers: {list(col_map.keys())}")
            return result

        for r in range(header_row + 1, ws.max_row + 1):
            symbol_raw = ws.cell(r, sym_col).value
            if not symbol_raw or str(symbol_raw).strip() == '':
                continue

            symbol = self._normalize_symbol(str(symbol_raw).strip().upper())
            if '-' in symbol and symbol.split('-')[0] in ('NSE', 'BSE'):
                symbol = symbol.split('-', 1)[-1]

            isin = str(ws.cell(r, isin_col).value or '') if isin_col else ''
            qty = _safe_float(ws.cell(r, qty_col).value)
            avg_price = _safe_float(ws.cell(r, avg_col).value) if avg_col else 0
            closing_price = _safe_float(ws.cell(r, ltp_col).value) if ltp_col else 0
            cur_value = _safe_float(ws.cell(r, val_col).value) if val_col else 0

            if qty <= 0:
                continue

            result.stocks.append(StockHolding(
                symbol=symbol,
                isin=isin,
                quantity=qty,
                avg_price=round(avg_price, 2),
                invested=round(qty * avg_price, 2),
                closing_price=round(closing_price, 2),
                closing_value=cur_value if cur_value > 0 else round(qty * closing_price, 2),
                broker=self.broker_name,
            ))

        return result
