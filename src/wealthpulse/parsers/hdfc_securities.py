"""
HDFC Securities broker statement parser.

Parses HDFC Securities equity holdings CSV/XLSX export.
Download: HDFC Securities Web → My Portfolio → Holdings → Download

Expected format:
  Stock Symbol / Scrip Code, Company Name, ISIN, Qty, Buy Avg, LTP, Current Value, P&L

Header row is auto-detected.
"""

import csv
from .base import BaseParser, ParseResult, StockHolding


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(',', '').replace('₹', '').strip())
    except (ValueError, TypeError):
        return 0.0


def _find_col(col_map: dict, candidates: list[str]) -> int | None:
    for c in candidates:
        if c in col_map:
            return col_map[c]
    return None


class HDFCSecuritiesParser(BaseParser):
    """Parser for HDFC Securities equity holdings export."""

    broker_name = "HDFC Securities"
    file_patterns = [
        "HDFC_*.xlsx", "hdfc_*.xlsx", "HDFCSec_*.xlsx",
        "HDFC_*.csv", "hdfc_*.csv", "HDFCSec_*.csv",
        "hdfcsec_*.xlsx", "hdfcsec_*.csv",
        "HDFCSecurities_*.xlsx", "HDFCSecurities_*.csv",
    ]

    def parse(self, filepath: str) -> ParseResult:
        if filepath.lower().endswith('.csv'):
            return self._parse_csv(filepath)
        return self._parse_xlsx(filepath)

    def _parse_csv(self, filepath: str) -> ParseResult:
        result = ParseResult(broker_name=self.broker_name)

        try:
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    result.errors.append("Empty CSV file")
                    return result

                headers = {h.strip().lower(): h for h in reader.fieldnames}

                for row in reader:
                    symbol_raw = None
                    for key in ['stock symbol', 'symbol', 'scrip code', 'scrip', 'stock code',
                                'nse symbol', 'trading symbol', 'security name']:
                        mapped = headers.get(key)
                        if mapped and row.get(mapped):
                            symbol_raw = row[mapped].strip()
                            break

                    if not symbol_raw:
                        continue

                    symbol = self._normalize_symbol(symbol_raw.upper())
                    if ':' in symbol:
                        symbol = symbol.split(':')[-1]

                    qty = 0
                    for key in ['qty', 'quantity', 'net qty', 'total qty', 'free qty']:
                        mapped = headers.get(key)
                        if mapped and row.get(mapped):
                            qty = _safe_float(row[mapped])
                            break

                    avg_price = 0
                    for key in ['buy avg', 'avg price', 'avg cost', 'average price', 'buy avg price']:
                        mapped = headers.get(key)
                        if mapped and row.get(mapped):
                            avg_price = _safe_float(row[mapped])
                            break

                    ltp = 0
                    for key in ['ltp', 'last price', 'current price', 'close price', 'mkt price']:
                        mapped = headers.get(key)
                        if mapped and row.get(mapped):
                            ltp = _safe_float(row[mapped])
                            break

                    isin = ''
                    for key in ['isin', 'isin code']:
                        mapped = headers.get(key)
                        if mapped and row.get(mapped):
                            isin = row[mapped].strip()
                            break

                    if qty <= 0:
                        continue

                    result.stocks.append(StockHolding(
                        symbol=symbol,
                        isin=isin,
                        quantity=qty,
                        avg_price=round(avg_price, 2),
                        invested=round(qty * avg_price, 2),
                        closing_price=round(ltp, 2),
                        closing_value=round(qty * ltp, 2),
                        broker=self.broker_name,
                    ))

        except Exception as e:
            result.errors.append(f"Error parsing HDFC Securities CSV: {e}")

        return result

    def _parse_xlsx(self, filepath: str) -> ParseResult:
        import openpyxl
        result = ParseResult(broker_name=self.broker_name)

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            result.errors.append(f"Cannot open {filepath}: {e}")
            return result

        ws = wb.active

        header_row = None
        col_map = {}
        for r in range(1, min(30, ws.max_row + 1)):
            for c in range(1, min(20, ws.max_column + 1)):
                val = str(ws.cell(r, c).value or '').strip().lower()
                if val in ('stock symbol', 'symbol', 'scrip code', 'scrip', 'stock code',
                           'nse symbol', 'trading symbol', 'security name'):
                    header_row = r
                    break
            if header_row:
                break

        if not header_row:
            result.errors.append("Could not find header row in HDFC Securities file")
            return result

        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(header_row, c).value or '').strip().lower()
            col_map[val] = c

        sym_col = _find_col(col_map, ['stock symbol', 'symbol', 'scrip code', 'scrip', 'nse symbol',
                                       'trading symbol', 'security name'])
        isin_col = _find_col(col_map, ['isin', 'isin code'])
        qty_col = _find_col(col_map, ['qty', 'quantity', 'net qty', 'total qty', 'free qty'])
        avg_col = _find_col(col_map, ['buy avg', 'avg price', 'avg cost', 'average price', 'buy avg price'])
        ltp_col = _find_col(col_map, ['ltp', 'last price', 'current price', 'close price'])
        val_col = _find_col(col_map, ['current value', 'mkt value', 'market value'])

        if not sym_col or not qty_col:
            result.errors.append(f"Cannot map columns. Headers: {list(col_map.keys())}")
            return result

        for r in range(header_row + 1, ws.max_row + 1):
            symbol_raw = ws.cell(r, sym_col).value
            if not symbol_raw or str(symbol_raw).strip() == '':
                continue

            symbol = self._normalize_symbol(str(symbol_raw).strip().upper())
            if ':' in symbol:
                symbol = symbol.split(':')[-1]

            isin = str(ws.cell(r, isin_col).value or '') if isin_col else ''
            qty = _safe_float(ws.cell(r, qty_col).value)
            avg_price = _safe_float(ws.cell(r, avg_col).value) if avg_col else 0
            closing_price = _safe_float(ws.cell(r, ltp_col).value) if ltp_col else 0
            cur_value = _safe_float(ws.cell(r, val_col).value) if val_col else 0

            if qty <= 0:
                continue

            result.stocks.append(StockHolding(
                symbol=symbol,
                isin=isin,
                quantity=qty,
                avg_price=round(avg_price, 2),
                invested=round(qty * avg_price, 2),
                closing_price=round(closing_price, 2),
                closing_value=cur_value if cur_value > 0 else round(qty * closing_price, 2),
                broker=self.broker_name,
            ))

        return result
