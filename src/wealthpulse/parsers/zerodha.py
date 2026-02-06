"""
Zerodha broker statement parser.

Parses Zerodha Console equity holdings XLSX export.
Expected format (data starts at col B):
  Symbol | ISIN | Sector | Qty Available | Qty Discrepant | Qty Long Term |
  Qty Pledged(Margin) | Qty Pledged(Loan) | Average Price | Previous Closing Price |
  Unrealized P&L | Unrealized P&L Pct
Header row is auto-detected (looks for "Symbol" in column B).
"""

from .base import BaseParser, ParseResult, StockHolding


class ZerodhaParser(BaseParser):
    """Parser for Zerodha Console equity holdings XLSX export."""

    broker_name = "Zerodha"
    file_patterns = ["Zerodha_*.xlsx", "zerodha_*.xlsx", "kite_*.xlsx"]

    def parse(self, filepath: str) -> ParseResult:
        import openpyxl
        result = ParseResult(broker_name=self.broker_name)

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            result.errors.append(f"Cannot open {filepath}: {e}")
            return result

        # Try 'Equity' sheet first, then first sheet
        if 'Equity' in wb.sheetnames:
            ws = wb['Equity']
        else:
            ws = wb[wb.sheetnames[0]]

        # Find header row with 'Symbol' (Zerodha has blank col A, data starts col B)
        header_row = None
        for r in range(1, min(35, ws.max_row + 1)):
            val = ws.cell(r, 2).value
            if val and 'symbol' in str(val).lower():
                header_row = r
                break

        if not header_row:
            result.errors.append("Could not find header row with 'Symbol' in column B")
            return result

        # Parse rows
        for r in range(header_row + 1, ws.max_row + 1):
            symbol = ws.cell(r, 2).value    # Col B
            if not symbol or str(symbol).strip() == '':
                continue

            symbol = self._normalize_symbol(str(symbol).strip())

            isin = str(ws.cell(r, 3).value or '')           # Col C
            sector = str(ws.cell(r, 4).value or '')          # Col D
            qty = float(ws.cell(r, 5).value or 0)            # Col E: Qty Available
            avg_price = float(ws.cell(r, 10).value or 0)     # Col J: Average Price
            closing_price = float(ws.cell(r, 11).value or 0) # Col K: Previous Closing Price
            invested = round(qty * avg_price, 2)
            closing_value = round(qty * closing_price, 2)

            result.stocks.append(StockHolding(
                symbol=symbol,
                isin=isin,
                quantity=qty,
                avg_price=round(avg_price, 2),
                invested=invested,
                closing_price=round(closing_price, 2),
                closing_value=round(closing_value, 2),
                broker=self.broker_name,
                sector=sector,
            ))

        return result
