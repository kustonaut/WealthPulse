"""
Groww broker statement parser.

Parses Groww equity holdings XLSX export.
Expected format: Stock Name | ISIN | Quantity | Avg buy price | Buy value | Closing price | Closing value | Unrealised P&L
Header row is auto-detected (looks for "Stock Name").
"""

import re
from .base import BaseParser, ParseResult, StockHolding

# Common Groww stock name → NSE symbol mapping
# Groww uses full company names, we need trading symbols
GROWW_NAME_MAP = {
    'Aarti Industries Limited': 'AARTIIND',
    'Aavas Financiers Limited': 'AAVAS',
    'Alkyl Amines Chemicals Limited': 'ALKYLAMINE',
    'Avenue Supermarts Limited': 'DMART',
    'Bajaj Finance Limited': 'BAJFINANCE',
    'Bajaj Housing Finance Limited': 'BAJAJHFL',
    'Bandhan Bank Limited': 'BANDHANBNK',
    'Bata India Limited': 'BATAINDIA',
    'Dr. Lal PathLabs Limited': 'LALPATHLAB',
    'Easy Trip Planners Limited': 'EASEMYTRIP',
    'GMM Pfaudler Limited': 'GMMPFAUDLR',
    'HDFC Bank Limited': 'HDFCBANK',
    'HDFC Life Insurance Company Limited': 'HDFCLIFE',
    'ICICI Lombard General Insurance Company Limited': 'ICICIGI',
    'Info Edge (India) Limited': 'NAUKRI',
    'Infosys Limited': 'INFOSYS',
    'Jio Financial Services Limited': 'JIOFIN',
    'Jubilant FoodWorks Limited': 'JUBLFOOD',
    'LTIMindtree Limited': 'LTIM',
    'Maruti Suzuki India Limited': 'MARUTI',
    'Mold-Tek Packaging Limited': 'MOLDTKPAC',
    'Nestle India Limited': 'NESTLEIND',
    'Pidilite Industries Limited': 'PIDILITIND',
    'Prince Pipes and Fittings Limited': 'PRINCEPIPE',
    'Reliance Industries Limited': 'RELIANCE',
    'SBI - ETF Nifty 50': 'SETFNIF50',
    'Tata Consultancy Services Limited': 'TCS',
    'Vodafone Idea Limited': 'IDEA',
}

# ISIN → Symbol fallback
ISIN_MAP = {
    'INE020B01018': 'RELIANCE',
    'INE467B01029': 'TCS',
    'INE009A01021': 'INFOSYS',
    'INE040A01034': 'HDFCBANK',
    'INE296A01024': 'BAJFINANCE',
    'INE795G01014': 'BAJAJHFL',
    'INE118H01025': 'BANDHANBNK',
    'INE176A01028': 'BATAINDIA',
    'INE600K01018': 'AAVAS',
    'INE498S01024': 'EASEMYTRIP',
    'INE541W01024': 'JIOFIN',
    'INE797F01020': 'JUBLFOOD',
    'INE214T01019': 'LTIM',
    'INE585B01010': 'MARUTI',
    'INE00S201012': 'DMART',
    'INE934A01020': 'NESTLEIND',
    'INE318A01026': 'HDFCLIFE',
    'INE765G01017': 'ICICIGI',
}


class GrowwParser(BaseParser):
    """Parser for Groww equity holdings XLSX export."""

    broker_name = "Groww"
    file_patterns = ["Grow_*.xlsx", "Groww_*.xlsx", "groww_*.xlsx"]

    def parse(self, filepath: str) -> ParseResult:
        import openpyxl
        result = ParseResult(broker_name=self.broker_name)

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            result.errors.append(f"Cannot open {filepath}: {e}")
            return result

        ws = wb.active

        # Find header row containing "Stock Name"
        header_row = None
        for r in range(1, min(20, ws.max_row + 1)):
            val = ws.cell(r, 1).value
            if val and 'stock name' in str(val).lower():
                header_row = r
                break

        if not header_row:
            result.errors.append("Could not find header row with 'Stock Name'")
            return result

        # Try to extract statement date
        for r in range(1, header_row):
            val = str(ws.cell(r, 1).value or '')
            if 'as on' in val.lower():
                m = re.search(r'(\d{2}-\d{2}-\d{4})', val)
                if m:
                    result.statement_date = m.group(1)
                break

        # Parse rows: Stock Name | ISIN | Qty | Avg Price | Buy Value | Closing Price | Closing Value
        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name or str(name).strip() == '':
                continue

            name = str(name).strip()
            isin = str(ws.cell(r, 2).value or '').strip()
            qty = float(ws.cell(r, 3).value or 0)
            avg_price = float(ws.cell(r, 4).value or 0)
            invested = float(ws.cell(r, 5).value or 0)
            closing_price = float(ws.cell(r, 6).value or 0)
            closing_value = float(ws.cell(r, 7).value or 0)

            # Resolve symbol from name or ISIN
            symbol = self._resolve_symbol(name, isin)

            result.stocks.append(StockHolding(
                symbol=symbol,
                isin=isin,
                quantity=qty,
                avg_price=round(avg_price, 2),
                invested=round(invested, 2),
                closing_price=round(closing_price, 2),
                closing_value=round(closing_value, 2),
                broker=self.broker_name,
            ))

        return result

    def _resolve_symbol(self, name: str, isin: str) -> str:
        """Resolve a Groww stock name to an NSE symbol."""
        # Try exact name match
        if name in GROWW_NAME_MAP:
            return GROWW_NAME_MAP[name]

        # Try ISIN match
        if isin in ISIN_MAP:
            return ISIN_MAP[isin]

        # Try fuzzy: first word match
        first_word = name.split()[0].upper() if name else ''
        for full_name, sym in GROWW_NAME_MAP.items():
            if full_name.split()[0].upper() == first_word:
                return sym

        # Fallback: use ISIN or cleaned name
        if isin and isin.startswith('INE'):
            return isin
        return re.sub(r'[^A-Z0-9]', '', name.upper())[:15]
